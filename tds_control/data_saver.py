import csv
import os
import queue
import threading
import time

import h5py
import numpy as np
from openpyxl import Workbook


DEFAULT_COLUMNS = [
    ("time", "time_stamp"),
    ("set_T", "set_temperature"),
    ("T", "measured_temperature"),
    ("h_f", "heat_flux"),
    ("V", "voltage"),
    ("I", "current"),
    ("C_V", "calculated_voltage"),
    ("P", "sample_power"),
    ("R_ohm", "calculated_resistance"),
]


class ExperimentDataSaver:
    """
    Persist experiment data in the background so instrument control never waits on disk IO.
    """

    def __init__(
        self,
        experiment_dir,
        r_vs_t,
        columns=None,
        flush_interval_s=5.0,
        batch_size=10,
        calibration_note=None,
    ):
        self.experiment_dir = experiment_dir
        self.r_vs_t = np.array(r_vs_t, dtype=float)
        self.calibration_note = str(calibration_note).strip() if calibration_note else None
        self.columns = list(columns or DEFAULT_COLUMNS)
        self.flush_interval_s = max(float(flush_interval_s), 0.5)
        self.batch_size = max(int(batch_size), 1)

        self.csv_path = os.path.join(self.experiment_dir, "data.csv")
        self.xlsx_path = os.path.join(self.experiment_dir, "data.xlsx")
        self.h5_path = os.path.join(self.experiment_dir, "data.h5")
        self.r_vs_t_path = os.path.join(self.experiment_dir, "r_vs_t.csv")
        self.r_vs_t_pdf_path = os.path.join(self.experiment_dir, "corrected_r_vs_t_curve.pdf")
        self.calibration_info_path = os.path.join(self.experiment_dir, "calibration_info.txt")

        self._queue = queue.Queue()
        self._stop_token = object()
        self._ready = threading.Event()
        self._thread = threading.Thread(target=self._worker, daemon=True, name="experiment-data-saver")
        self._closed = False
        self._error = None
        self.rows_written = 0

    @property
    def error(self):
        return self._error

    def start(self):
        os.makedirs(self.experiment_dir, exist_ok=True)
        self._write_r_vs_t_snapshot()
        self._write_r_vs_t_pdf()
        self._write_calibration_info()
        self._thread.start()
        self._ready.wait(timeout=10.0)
        self.raise_if_error()
        return self

    def enqueue(self, row):
        self.raise_if_error()
        if self._closed:
            raise RuntimeError("Experiment data saver is already closed.")
        if len(row) != len(self.columns):
            raise ValueError(
                f"Expected {len(self.columns)} saved values, received {len(row)}."
            )
        self._queue.put(tuple(float(value) for value in row))

    def finalize(self, timeout=15.0):
        if self._closed:
            self.raise_if_error()
            return
        self._closed = True
        self._queue.put(self._stop_token)
        self._thread.join(timeout=timeout)
        if self._thread.is_alive():
            raise RuntimeError("Timed out while waiting for experiment data to finish saving.")
        self.raise_if_error()

    def raise_if_error(self):
        if self._error is not None:
            raise RuntimeError(f"Background data saver failed: {self._error}")

    def _write_r_vs_t_snapshot(self):
        with open(self.r_vs_t_path, "w", newline="", encoding="utf-8") as csv_file:
            writer = csv.writer(csv_file)
            writer.writerow(["resistivity", "temperature"])
            for resistivity, temperature in self.r_vs_t.T:
                writer.writerow([float(resistivity), float(temperature)])

    def _ordered_finite_curve(self):
        if self.r_vs_t.ndim != 2 or self.r_vs_t.shape[0] != 2:
            raise ValueError("The corrected R-vs-T curve must have shape (2, N).")
        resistance = np.asarray(self.r_vs_t[0, :], dtype=float)
        temperature = np.asarray(self.r_vs_t[1, :], dtype=float)
        finite = np.isfinite(resistance) & np.isfinite(temperature)
        if int(np.count_nonzero(finite)) < 2:
            raise ValueError("The corrected R-vs-T curve needs at least two finite points.")
        resistance = resistance[finite]
        temperature = temperature[finite]
        order = np.argsort(temperature)
        return resistance[order], temperature[order]

    @staticmethod
    def _escape_pdf_text(value):
        return (
            str(value)
            .replace("\\", "\\\\")
            .replace("(", "\\(")
            .replace(")", "\\)")
        )

    @classmethod
    def _pdf_text_command(
        cls,
        text,
        x,
        y,
        size,
        *,
        bold=False,
        align="left",
        rotated=False,
    ):
        escaped_text = cls._escape_pdf_text(text)
        estimated_width = len(str(text)) * float(size) * 0.52
        if rotated:
            if align == "center":
                y -= estimated_width / 2.0
            elif align == "right":
                y -= estimated_width
            matrix = f"0 1 -1 0 {float(x):.2f} {float(y):.2f} Tm"
        else:
            if align == "center":
                x -= estimated_width / 2.0
            elif align == "right":
                x -= estimated_width
            matrix = f"1 0 0 1 {float(x):.2f} {float(y):.2f} Tm"
        font_name = "F2" if bold else "F1"
        return (
            f"BT /{font_name} {float(size):.2f} Tf "
            f"{matrix} ({escaped_text}) Tj ET"
        )

    @staticmethod
    def _write_pdf_document(path, content):
        content_bytes = content.encode("ascii")
        objects = [
            b"<< /Type /Catalog /Pages 2 0 R >>",
            b"<< /Type /Pages /Kids [3 0 R] /Count 1 >>",
            (
                b"<< /Type /Page /Parent 2 0 R "
                b"/MediaBox [0 0 595.28 841.89] "
                b"/Resources << /Font << /F1 5 0 R /F2 6 0 R >> >> "
                b"/Contents 4 0 R >>"
            ),
            (
                f"<< /Length {len(content_bytes)} >>\nstream\n".encode("ascii")
                + content_bytes
                + b"\nendstream"
            ),
            b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>",
            b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica-Bold >>",
            (
                b"<< /Title (Corrected R vs. T curve used by the experiment) "
                b"/Creator (TDS Control) >>"
            ),
        ]

        pdf_data = bytearray(b"%PDF-1.4\n%\xe2\xe3\xcf\xd3\n")
        offsets = [0]
        for object_number, object_data in enumerate(objects, start=1):
            offsets.append(len(pdf_data))
            pdf_data.extend(f"{object_number} 0 obj\n".encode("ascii"))
            pdf_data.extend(object_data)
            pdf_data.extend(b"\nendobj\n")

        xref_offset = len(pdf_data)
        pdf_data.extend(f"xref\n0 {len(objects) + 1}\n".encode("ascii"))
        pdf_data.extend(b"0000000000 65535 f \n")
        for offset in offsets[1:]:
            pdf_data.extend(f"{offset:010d} 00000 n \n".encode("ascii"))
        pdf_data.extend(
            (
                f"trailer\n<< /Size {len(objects) + 1} /Root 1 0 R /Info 7 0 R >>\n"
                f"startxref\n{xref_offset}\n%%EOF\n"
            ).encode("ascii")
        )

        with open(path, "wb") as pdf_file:
            pdf_file.write(pdf_data)

    def _write_r_vs_t_pdf(self):
        resistance, temperature = self._ordered_finite_curve()

        page_width = 595.28
        page_height = 841.89
        plot_left = 85.0
        plot_bottom = 145.0
        plot_width = 460.0
        plot_height = 580.0

        x_min = float(np.min(temperature))
        x_max = float(np.max(temperature))
        y_min = float(np.min(resistance))
        y_max = float(np.max(resistance))
        x_span = max(x_max - x_min, 1e-12)
        y_span = max(y_max - y_min, 1e-12)
        x_min -= 0.03 * x_span
        x_max += 0.03 * x_span
        y_min -= 0.05 * y_span
        y_max += 0.05 * y_span

        commands = [
            "1 1 1 rg 0 0 595.28 841.89 re f",
            "0 g",
            self._pdf_text_command(
                "Corrected R vs. T Curve Used by the Experiment",
                page_width / 2.0,
                page_height - 48.0,
                16.0,
                bold=True,
                align="center",
            ),
        ]

        tick_count = 6
        for tick_index in range(tick_count):
            fraction = tick_index / (tick_count - 1)
            x_position = plot_left + fraction * plot_width
            y_position = plot_bottom + fraction * plot_height
            x_value = x_min + fraction * (x_max - x_min)
            y_value = y_min + fraction * (y_max - y_min)

            commands.extend(
                [
                    (
                        "0.85 G 0.5 w "
                        f"{x_position:.2f} {plot_bottom:.2f} m "
                        f"{x_position:.2f} {plot_bottom + plot_height:.2f} l S"
                    ),
                    (
                        "0.85 G 0.5 w "
                        f"{plot_left:.2f} {y_position:.2f} m "
                        f"{plot_left + plot_width:.2f} {y_position:.2f} l S"
                    ),
                    "0 g",
                    self._pdf_text_command(
                        f"{x_value:.3g}",
                        x_position,
                        plot_bottom - 20.0,
                        8.0,
                        align="center",
                    ),
                    self._pdf_text_command(
                        f"{y_value:.5g}",
                        plot_left - 8.0,
                        y_position - 3.0,
                        8.0,
                        align="right",
                    ),
                ]
            )

        commands.append(
            (
                "0 G 1.2 w "
                f"{plot_left:.2f} {plot_bottom:.2f} "
                f"{plot_width:.2f} {plot_height:.2f} re S"
            )
        )

        curve_commands = ["0.08 0.39 0.71 RG 1.6 w"]
        for point_index, (resistance_value, temperature_value) in enumerate(
            zip(resistance, temperature)
        ):
            x_position = plot_left + (
                (float(temperature_value) - x_min) / (x_max - x_min)
            ) * plot_width
            y_position = plot_bottom + (
                (float(resistance_value) - y_min) / (y_max - y_min)
            ) * plot_height
            operator = "m" if point_index == 0 else "l"
            curve_commands.append(f"{x_position:.2f} {y_position:.2f} {operator}")
        curve_commands.append("S")
        commands.append("\n".join(curve_commands))

        commands.extend(
            [
                "0 g",
                self._pdf_text_command(
                    "Temperature (deg C)",
                    plot_left + plot_width / 2.0,
                    plot_bottom - 48.0,
                    11.0,
                    bold=True,
                    align="center",
                ),
                self._pdf_text_command(
                    "Corrected sample resistance (Ohm)",
                    28.0,
                    plot_bottom + plot_height / 2.0,
                    11.0,
                    bold=True,
                    align="center",
                    rotated=True,
                ),
                self._pdf_text_command(
                    "Exact corrected curve supplied to the experiment after T0 scaling.",
                    page_width / 2.0,
                    66.0,
                    8.5,
                    align="center",
                ),
                self._pdf_text_command(
                    "Vertical axis is Kelvin resistance (V/I) in ohms, not bulk resistivity in ohm-metre.",
                    page_width / 2.0,
                    49.0,
                    8.5,
                    align="center",
                ),
            ]
        )

        self._write_pdf_document(
            self.r_vs_t_pdf_path,
            "\n".join(commands),
        )

    def _write_calibration_info(self):
        if not self.calibration_note:
            return
        with open(self.calibration_info_path, "w", encoding="utf-8") as info_file:
            info_file.write("T0 calibration warning\n")
            info_file.write(self.calibration_note)
            info_file.write("\n")

    def _create_h5_datasets(self, h5_file):
        datasets = {}
        for _, dataset_name in self.columns:
            datasets[dataset_name] = h5_file.create_dataset(
                dataset_name,
                shape=(0,),
                maxshape=(None,),
                dtype="f8",
                chunks=True,
            )
        if "calculated_resistance" in datasets:
            # Compatibility name requested by the GUI/user terminology. This is an
            # HDF5 hard link to the same V/I data, so it does not duplicate storage.
            h5_file["calculated_resistivity"] = datasets["calculated_resistance"]
        return datasets

    def _create_excel_workbook(self):
        workbook = Workbook(write_only=True)
        data_sheet = workbook.create_sheet("Experiment Data")
        data_sheet.append([column_name for column_name, _ in self.columns])

        curve_sheet = workbook.create_sheet("Corrected R vs T")
        curve_sheet.append(["temperature_C", "corrected_resistance_ohm"])
        resistance, temperature = self._ordered_finite_curve()
        for resistance_value, temperature_value in zip(resistance, temperature):
            curve_sheet.append([float(temperature_value), float(resistance_value)])
        return workbook, data_sheet

    def _append_batch(self, batch, csv_writer, csv_file, datasets, h5_file, excel_sheet):
        if not batch:
            return

        for row in batch:
            csv_writer.writerow(row)
            excel_sheet.append(row)
        csv_file.flush()

        array_batch = np.asarray(batch, dtype=float)
        start_index = self.rows_written
        end_index = start_index + len(array_batch)
        for column_index, (_, dataset_name) in enumerate(self.columns):
            dataset = datasets[dataset_name]
            dataset.resize((end_index,))
            dataset[start_index:end_index] = array_batch[:, column_index]
        h5_file.flush()
        self.rows_written = end_index

    def _worker(self):
        batch = []
        last_flush = time.time()
        try:
            workbook, excel_sheet = self._create_excel_workbook()
            with open(self.csv_path, "w", newline="", encoding="utf-8") as csv_file, h5py.File(
                self.h5_path, "w"
            ) as h5_file:
                csv_writer = csv.writer(csv_file)
                csv_writer.writerow([column_name for column_name, _ in self.columns])
                datasets = self._create_h5_datasets(h5_file)
                self._ready.set()

                while True:
                    timeout = max(0.1, self.flush_interval_s - (time.time() - last_flush))
                    try:
                        item = self._queue.get(timeout=timeout)
                    except queue.Empty:
                        item = None

                    if item is self._stop_token:
                        self._append_batch(
                            batch,
                            csv_writer,
                            csv_file,
                            datasets,
                            h5_file,
                            excel_sheet,
                        )
                        break

                    if item is not None:
                        batch.append(item)

                    if batch and (
                        len(batch) >= self.batch_size
                        or time.time() - last_flush >= self.flush_interval_s
                    ):
                        self._append_batch(
                            batch,
                            csv_writer,
                            csv_file,
                            datasets,
                            h5_file,
                            excel_sheet,
                        )
                        batch = []
                        last_flush = time.time()

            workbook.save(self.xlsx_path)
        except Exception as exc:
            self._error = exc
            self._ready.set()
