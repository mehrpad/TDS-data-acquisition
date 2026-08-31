import csv
import tempfile
import unittest
from pathlib import Path

import h5py
import numpy as np
from openpyxl import load_workbook

from tds_control.data_saver import ExperimentDataSaver


class ExperimentDataSaverOutputTests(unittest.TestCase):
    def test_writes_excel_resistance_datasets_and_corrected_curve_pdf(self):
        corrected_curve = np.array(
            [
                [10.0, 11.5, 14.0],
                [23.0, 100.0, 250.0],
            ],
            dtype=float,
        )
        saved_row = [
            1_700_000_000.0,
            50.0,
            48.5,
            0.0,
            0.2,
            0.01,
            0.25,
            0.002,
            20.0,
        ]

        with tempfile.TemporaryDirectory() as temporary_directory:
            result_directory = Path(temporary_directory)
            saver = ExperimentDataSaver(
                result_directory,
                corrected_curve,
                flush_interval_s=0.5,
                batch_size=1,
            ).start()
            saver.enqueue(saved_row)
            saver.finalize()

            expected_files = {
                "data.csv",
                "data.xlsx",
                "data.h5",
                "r_vs_t.csv",
                "corrected_r_vs_t_curve.pdf",
            }
            self.assertTrue(
                expected_files.issubset({path.name for path in result_directory.iterdir()})
            )

            with (result_directory / "data.csv").open(newline="", encoding="utf-8") as csv_file:
                csv_rows = list(csv.reader(csv_file))
            self.assertEqual(csv_rows[0][-1], "R_ohm")
            self.assertAlmostEqual(float(csv_rows[1][-1]), 20.0)

            workbook = load_workbook(result_directory / "data.xlsx", read_only=True, data_only=True)
            try:
                data_rows = list(workbook["Experiment Data"].iter_rows(values_only=True))
                self.assertEqual(data_rows[0][-1], "R_ohm")
                self.assertAlmostEqual(float(data_rows[1][-1]), 20.0)

                curve_rows = list(workbook["Corrected R vs T"].iter_rows(values_only=True))
                self.assertEqual(
                    curve_rows[0],
                    ("temperature_C", "corrected_resistance_ohm"),
                )
                self.assertEqual(curve_rows[1], (23.0, 10.0))
                self.assertEqual(curve_rows[-1], (250.0, 14.0))
            finally:
                workbook.close()

            with h5py.File(result_directory / "data.h5", "r") as h5_file:
                self.assertAlmostEqual(float(h5_file["calculated_resistance"][0]), 20.0)
                self.assertAlmostEqual(float(h5_file["calculated_resistivity"][0]), 20.0)
                self.assertEqual(
                    h5_file["calculated_resistance"].id,
                    h5_file["calculated_resistivity"].id,
                )

            pdf_path = result_directory / "corrected_r_vs_t_curve.pdf"
            self.assertGreater(pdf_path.stat().st_size, 1_000)
            with pdf_path.open("rb") as pdf_file:
                self.assertEqual(pdf_file.read(4), b"%PDF")


if __name__ == "__main__":
    unittest.main()
